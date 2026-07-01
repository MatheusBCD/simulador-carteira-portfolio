# -*- coding: utf-8 -*-
"""
gerar_excel.py
Gera o arquivo Simulador_Cliente.xlsx: uma planilha onde o cliente preenche
os dados (células azuis) e todos os resultados são calculados automaticamente
por fórmulas do próprio Excel (sem precisar rodar Python depois de pronto).

Para gerar o arquivo:
    python gerar_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

import mercado

print("Atualizando premissas de mercado (Selic, IPCA, Ibovespa)...")
mercado.atualizar_premissas()

AZUL = Font(color="0000FF")
PRETO = Font(color="000000")
VERDE = Font(color="008000")
NEGRITO = Font(bold=True)
TITULO = Font(bold=True, size=14)
SUBTITULO = Font(bold=True, size=11)
AMARELO = PatternFill("solid", start_color="FFFF00")
CINZA = PatternFill("solid", start_color="D9D9D9")
CENTRO = Alignment(horizontal="center", vertical="center")
PRAZO_MAX_ANOS = 30

wb = Workbook()

# ======================================================================
# SHEET: Instruções
# ======================================================================
ins = wb.active
ins.title = "Instrucoes"
ins["A1"] = "Simulador de Perfil do Cliente — Projeção de Resultados"
ins["A1"].font = TITULO
ins["A3"] = "Como usar esta planilha:"
ins["A3"].font = SUBTITULO
passos = [
    "1. Vá até a aba 'Perfil_Risco' e responda as 4 perguntas (células azuis com lista suspensa).",
    "2. Vá até a aba 'Dados_Cliente' e preencha os dados em azul (nome, idade, renda, valores e prazo).",
    "3. Vá até a aba 'Resultado' para ver a projeção final, os indicadores e os gráficos.",
    "4. Se quiser, ajuste as premissas de mercado (retorno/risco/inflação) na aba 'Mercado'.",
    "",
    "Células AZUIS = você pode editar.  Células PRETAS = calculadas automaticamente, não edite.",
    f"Prazo máximo suportado por esta planilha: {PRAZO_MAX_ANOS} anos.",
]
for i, linha in enumerate(passos, start=4):
    ins[f"A{i}"] = linha
ins.column_dimensions["A"].width = 100

# ======================================================================
# SHEET: Mercado (premissas)
# ======================================================================
mkt = wb.create_sheet("Mercado")
mkt["A1"] = "Premissas de Mercado (edite os valores em azul)"
mkt["A1"].font = TITULO

headers = ["Código", "Classe de Ativo", "Retorno Anual Esperado", "Volatilidade Anual", "Fonte / Observação"]
for c, h in enumerate(headers, start=1):
    cel = mkt.cell(row=3, column=c, value=h)
    cel.font = NEGRITO
    cel.fill = CINZA

classes = [
    (codigo, dados["nome"], dados["retorno_anual"], dados["volatilidade_anual"])
    for codigo, dados in mercado.CLASSES_ATIVOS.items()
]
for i, (codigo, nome, ret, vol) in enumerate(classes):
    r = 4 + i
    mkt.cell(row=r, column=1, value=codigo)
    mkt.cell(row=r, column=2, value=nome)
    c_ret = mkt.cell(row=r, column=3, value=ret)
    c_ret.font = AZUL
    c_ret.number_format = "0.0%"
    c_vol = mkt.cell(row=r, column=4, value=vol)
    c_vol.font = AZUL
    c_vol.number_format = "0.0%"
    mkt.cell(row=r, column=5, value="Estimativa própria — ajustar conforme fonte oficial").font = Font(italic=True, size=9)

mkt["A11"] = "Inflação anual esperada (IPCA)"
mkt["B11"] = mercado.INFLACAO_ANUAL
mkt["B11"].font = AZUL
mkt["B11"].number_format = "0.0%"
mkt["A12"] = "Taxa livre de risco (CDI / Selic)"
mkt["B12"] = mercado.TAXA_LIVRE_RISCO_ANUAL
mkt["B12"].font = AZUL
mkt["B12"].number_format = "0.0%"

mkt["A13"] = "Última atualização com dados reais"
status_atualizacao = mercado.DATA_ULTIMA_ATUALIZACAO or "NÃO ATUALIZADO (usando valores de referência)"
mkt["B13"] = status_atualizacao
mkt["B13"].font = Font(italic=True, size=9, color="808080")

mkt["A14"] = ("Fonte sugerida para atualização: Banco Central do Brasil (bcb.gov.br) para "
              "CDI/Selic, e IBGE (ibge.gov.br/explica/inflacao.php) para o IPCA.")
mkt["A14"].font = Font(italic=True, size=9)
for col, w in zip("ABCDE", [20, 42, 22, 20, 45]):
    mkt.column_dimensions[col].width = w

# ======================================================================
# SHEET: Perfil_Risco (questionário)
# ======================================================================
qr = wb.create_sheet("Perfil_Risco")
qr["A1"] = "Questionário de Perfil de Risco (Suitability)"
qr["A1"].font = TITULO

qr["A3"] = "Pergunta"
qr["B3"] = "Sua resposta (selecione)"
qr["C3"] = "Pontuação"
for c in "ABC":
    qr[f"{c}3"].font = NEGRITO
    qr[f"{c}3"].fill = CINZA

perguntas = [
    ("Qual seu objetivo principal com os investimentos?", [
        ("Preservar o capital, evitando perdas", 1),
        ("Obter renda com baixo risco", 2),
        ("Crescimento do patrimônio no médio prazo", 3),
        ("Maximizar retorno, aceitando oscilações", 4),
    ]),
    ("Qual seu horizonte de investimento?", [
        ("Até 1 ano", 1),
        ("De 1 a 3 anos", 2),
        ("De 3 a 7 anos", 3),
        ("Mais de 7 anos", 4),
    ]),
    ("Se sua carteira caísse 15% em um mês, você:", [
        ("Resgataria tudo imediatamente", 1),
        ("Resgataria parte do valor", 2),
        ("Manteria o investimento", 3),
        ("Aproveitaria para investir mais", 4),
    ]),
    ("Qual sua experiência com investimentos de risco (ações, fundos)?", [
        ("Nenhuma experiência", 1),
        ("Pouca experiência", 2),
        ("Experiência moderada", 3),
        ("Bastante experiência", 4),
    ]),
]

LISTA_COL_START = 8  # coluna H em diante, tabelas de apoio para as listas suspensas
linha_lista = 2
faixas_lookup = []
for pergunta, opcoes in perguntas:
    inicio = linha_lista
    for texto, pontos in opcoes:
        qr.cell(row=linha_lista, column=LISTA_COL_START, value=texto)
        qr.cell(row=linha_lista, column=LISTA_COL_START + 1, value=pontos)
        linha_lista += 1
    fim = linha_lista - 1
    faixas_lookup.append((inicio, fim))
    linha_lista += 1  # linha em branco entre listas

for i, ((pergunta, opcoes), (ini, fim)) in enumerate(zip(perguntas, faixas_lookup)):
    r = 4 + i
    qr.cell(row=r, column=1, value=pergunta)
    resp_cel = qr.cell(row=r, column=2, value=opcoes[0][0])
    resp_cel.font = AZUL
    resp_cel.fill = AMARELO
    col_lista = get_column_letter(LISTA_COL_START)
    faixa = f"$H${ini}:$H${fim}"
    dv = DataValidation(type="list", formula1=f"={faixa}", allow_blank=False)
    qr.add_data_validation(dv)
    dv.add(resp_cel)
    faixa_lookup = f"$H${ini}:$I${fim}"
    pont_cel = qr.cell(row=r, column=3, value=f'=IFERROR(VLOOKUP(B{r},{faixa_lookup},2,FALSE),0)')
    pont_cel.font = PRETO

qr["A9"] = "Pontuação total"
qr["A9"].font = NEGRITO
qr["B9"] = "=SUM(C4:C7)"
qr["B9"].font = PRETO

qr["A10"] = "Perfil identificado"
qr["A10"].font = NEGRITO
qr["B10"] = '=IF(B9<=7,"conservador",IF(B9<=12,"moderado","arrojado"))'
qr["B10"].font = PRETO
qr["B10"].fill = AMARELO

qr.column_dimensions["A"].width = 45
qr.column_dimensions["B"].width = 40
qr.column_dimensions["H"].width = 40
for col, w in zip(["C", "I"], [12, 12]):
    qr.column_dimensions[col].width = w

# ======================================================================
# SHEET: Alocacao_Modelo
# ======================================================================
am = wb.create_sheet("Alocacao_Modelo")
am["A1"] = "Alocação-modelo por Perfil de Risco (% da carteira)"
am["A1"].font = TITULO

codigos = [c[0] for c in classes]
am.cell(row=3, column=1, value="Perfil").font = NEGRITO
for c, codigo in enumerate(codigos, start=2):
    cel = am.cell(row=3, column=c, value=codigo)
    cel.font = NEGRITO
    cel.fill = CINZA

alocacoes = {
    "conservador": [0.60, 0.30, 0.10, 0.00, 0.00, 0.00],
    "moderado":    [0.35, 0.20, 0.20, 0.10, 0.00, 0.15],
    "arrojado":    [0.15, 0.00, 0.20, 0.30, 0.20, 0.15],
}
for i, (perfil, pesos) in enumerate(alocacoes.items()):
    r = 4 + i
    am.cell(row=r, column=1, value=perfil)
    for c, peso in enumerate(pesos, start=2):
        cel = am.cell(row=r, column=c, value=peso)
        cel.font = AZUL
        cel.number_format = "0%"

for col in "ABCDEFG":
    am.column_dimensions[col].width = 16

# ======================================================================
# SHEET: Dados_Cliente
# ======================================================================
dc = wb.create_sheet("Dados_Cliente")
dc["A1"] = "Dados do Cliente"
dc["A1"].font = TITULO

campos = [
    ("Nome", "Cliente"),
    ("Idade", 35),
    ("Renda mensal (R$)", 8000),
    ("Valor inicial disponível (R$)", 10000),
    ("Aporte mensal planejado (R$)", 1000),
    (f"Prazo do investimento (anos, máx. {PRAZO_MAX_ANOS})", 10),
]
for i, (rotulo, valor) in enumerate(campos):
    r = 3 + i
    dc.cell(row=r, column=1, value=rotulo).font = NEGRITO
    cel = dc.cell(row=r, column=2, value=valor)
    cel.font = AZUL
    cel.fill = AMARELO
    if "R$" in rotulo:
        cel.number_format = "R$ #,##0.00"

dv_prazo = DataValidation(type="whole", operator="between", formula1=1, formula2=PRAZO_MAX_ANOS,
                           allow_blank=False, showErrorMessage=True,
                           errorTitle="Prazo inválido",
                           error=f"Digite um número inteiro entre 1 e {PRAZO_MAX_ANOS} anos.")
dc.add_data_validation(dv_prazo)
dv_prazo.add(dc["B8"])

dc["A10"] = "Perfil de risco (do questionário)"
dc["A10"].font = NEGRITO
dc["B10"] = "=Perfil_Risco!B10"
dc["B10"].font = VERDE
dc["B10"].fill = AMARELO

dc["A12"] = "Alocação atual da carteira (conforme perfil)"
dc["A12"].font = SUBTITULO
dc["A13"] = "Classe de Ativo"
dc["B13"] = "Peso (%)"
dc["A13"].font = NEGRITO
dc["B13"].font = NEGRITO

for i in range(6):
    r = 14 + i
    dc.cell(row=r, column=1, value=f"=Mercado!B{4+i}").font = VERDE
    formula_peso = (f"=INDEX(Alocacao_Modelo!$B$4:$G$6,"
                     f"MATCH($B$10,Alocacao_Modelo!$A$4:$A$6,0),{i+1})")
    cel = dc.cell(row=r, column=2, value=formula_peso)
    cel.font = PRETO
    cel.number_format = "0.0%"

dc["A21"] = "Retorno anual estimado da carteira"
dc["A21"].font = NEGRITO
dc["B21"] = "=SUMPRODUCT(B14:B19,Mercado!C4:C9)"
dc["B21"].font = PRETO
dc["B21"].number_format = "0.00%"

dc["A22"] = "Volatilidade anual estimada da carteira"
dc["A22"].font = NEGRITO
dc["B22"] = "=SUMPRODUCT(B14:B19,Mercado!D4:D9)"
dc["B22"].font = PRETO
dc["B22"].number_format = "0.00%"

dc["A24"] = "Taxa mensal — cenário esperado"
dc["B24"] = "=(1+B21)^(1/12)-1"
dc["A25"] = "Retorno anual — cenário otimista (aprox. percentil 90)"
dc["B25"] = "=B21+1.2816*B22"
dc["A26"] = "Retorno anual — cenário pessimista (aprox. percentil 10, piso -30%)"
dc["B26"] = "=MAX(B21-1.2816*B22,-0.3)"
dc["A27"] = "Taxa mensal — cenário otimista"
dc["B27"] = "=(1+B25)^(1/12)-1"
dc["A28"] = "Taxa mensal — cenário pessimista"
dc["B28"] = "=(1+B26)^(1/12)-1"
for r in range(24, 29):
    dc.cell(row=r, column=1).font = Font(size=9, italic=True)
    cel = dc.cell(row=r, column=2)
    cel.font = PRETO
    cel.number_format = "0.000%"

dc.column_dimensions["A"].width = 48
dc.column_dimensions["B"].width = 22

# ======================================================================
# SHEET: Projecao (mês a mês) — motor de cálculo, 30 anos = 360 meses
# ======================================================================
pj = wb.create_sheet("Projecao")
pj["A1"] = "Mês"
pj["B1"] = "Ano"
pj["C1"] = "Total Aportado"
pj["D1"] = "Saldo — Cenário Esperado"
pj["E1"] = "Saldo — Cenário Otimista"
pj["F1"] = "Saldo — Cenário Pessimista"
for c in "ABCDEF":
    pj[f"{c}1"].font = NEGRITO
    pj[f"{c}1"].fill = CINZA

# Mês 0 (ponto de partida)
pj["A2"] = 0
pj["B2"] = 0
pj["C2"] = "=Dados_Cliente!B6"
pj["D2"] = "=Dados_Cliente!B6"
pj["E2"] = "=Dados_Cliente!B6"
pj["F2"] = "=Dados_Cliente!B6"

MESES = PRAZO_MAX_ANOS * 12
for m in range(1, MESES + 1):
    r = m + 2
    r_ant = r - 1
    pj.cell(row=r, column=1, value=f"=A{r_ant}+1")
    pj.cell(row=r, column=2, value=f"=A{r}/12")
    cond = f"$A{r}<=Dados_Cliente!$B$8*12"
    pj.cell(row=r, column=3, value=f'=IF({cond},C{r_ant}+Dados_Cliente!$B$7,C{r_ant})')
    pj.cell(row=r, column=4, value=f'=IF({cond},D{r_ant}*(1+Dados_Cliente!$B$24)+Dados_Cliente!$B$7,D{r_ant})')
    pj.cell(row=r, column=5, value=f'=IF({cond},E{r_ant}*(1+Dados_Cliente!$B$27)+Dados_Cliente!$B$7,E{r_ant})')
    pj.cell(row=r, column=6, value=f'=IF({cond},F{r_ant}*(1+Dados_Cliente!$B$28)+Dados_Cliente!$B$7,F{r_ant})')

for col in "CDEF":
    for r in range(2, MESES + 3):
        pj[f"{col}{r}"].number_format = "R$ #,##0"

for col, w in zip("ABCDEF", [8, 8, 18, 20, 20, 20]):
    pj.column_dimensions[col].width = w

pj.sheet_view.showGridLines = False
pj.sheet_state = "visible"

# ======================================================================
# SHEET: Resultado (dashboard final)
# ======================================================================
res = wb.create_sheet("Resultado")
res["A1"] = "Resultado da Projeção"
res["A1"].font = TITULO

linha_prazo_meses = "Dados_Cliente!$B$8*12"
match_formula = f"MATCH({linha_prazo_meses},Projecao!$A:$A,0)"

resumo = [
    ("Cliente", "=Dados_Cliente!B3", None),
    ("Perfil de risco", "=Dados_Cliente!B10", None),
    ("Retorno anual estimado da carteira", "=Dados_Cliente!B21", "0.00%"),
    ("Prazo (anos)", "=Dados_Cliente!B8", None),
    ("Total investido (aportes + inicial)", f"=INDEX(Projecao!C:C,{match_formula})", "R$ #,##0.00"),
]
for i, (rotulo, formula, fmt) in enumerate(resumo):
    r = 3 + i
    res.cell(row=r, column=1, value=rotulo).font = NEGRITO
    cel = res.cell(row=r, column=2, value=formula)
    cel.font = PRETO
    if fmt:
        cel.number_format = fmt

res["A9"] = "Cenários de Projeção (percentis 10 / 50 / 90 sobre o retorno da carteira)"
res["A9"].font = SUBTITULO

cenarios = [
    ("Cenário otimista", f"=INDEX(Projecao!E:E,{match_formula})"),
    ("Cenário esperado", f"=INDEX(Projecao!D:D,{match_formula})"),
    ("Cenário pessimista", f"=INDEX(Projecao!F:F,{match_formula})"),
]
for i, (rotulo, formula) in enumerate(cenarios):
    r = 10 + i
    res.cell(row=r, column=1, value=rotulo + ":").font = NEGRITO
    cel = res.cell(row=r, column=2, value=formula)
    cel.font = PRETO
    cel.number_format = "R$ #,##0.00"
    cel.fill = AMARELO

res["A14"] = "Indicadores"
res["A14"].font = SUBTITULO
res["A15"] = "Rendimento projetado (cenário esperado)"
res["B15"] = "=B11-B7"
res["A16"] = "CAGR (retorno anual composto, cenário esperado)"
res["B16"] = "=(B11/B7)^(1/B6)-1"
res["A17"] = "CAGR real (descontada a inflação)"
res["B17"] = "=(1+B16)/(1+Mercado!B11)-1"
for r, fmt in [(15, "R$ #,##0.00"), (16, "0.00%"), (17, "0.00%")]:
    res.cell(row=r, column=1).font = NEGRITO
    res.cell(row=r, column=2).font = PRETO
    res.cell(row=r, column=2).number_format = fmt

res.column_dimensions["A"].width = 42
res.column_dimensions["B"].width = 22

# --- Gráfico de linha: evolução dos 3 cenários ---
graf_linha = LineChart()
graf_linha.title = "Evolução do Patrimônio — Cenários"
graf_linha.style = 2
graf_linha.x_axis.title = "Ano"
graf_linha.y_axis.title = "Valor (R$)"
graf_linha.height = 9
graf_linha.width = 20

cats = Reference(pj, min_col=2, min_row=2, max_row=MESES + 2)
for col, nome in [(5, "Otimista"), (4, "Esperado"), (6, "Pessimista")]:
    dados = Reference(pj, min_col=col, min_row=1, max_row=MESES + 2)
    graf_linha.add_data(dados, titles_from_data=True)
graf_linha.set_categories(cats)
res.add_chart(graf_linha, "D3")
res["A28"] = ("Nota: o gráfico exibe até 30 anos; após o prazo escolhido, as linhas ficam "
              "estáveis (o valor não é mais projetado, é apenas mantido para referência visual).")
res["A28"].font = Font(italic=True, size=9)

# --- Gráfico de pizza: alocação da carteira ---
graf_pizza = PieChart()
graf_pizza.title = "Alocação Sugerida da Carteira"
graf_pizza.height = 9
graf_pizza.width = 12
dados_pizza = Reference(dc, min_col=2, min_row=14, max_row=19)
labels_pizza = Reference(dc, min_col=1, min_row=14, max_row=19)
graf_pizza.add_data(dados_pizza, titles_from_data=False)
graf_pizza.set_categories(labels_pizza)
res.add_chart(graf_pizza, "D22")

wb.save("/home/claude/simulador_cliente2/Simulador_Cliente.xlsx")
print("Arquivo Simulador_Cliente.xlsx gerado com sucesso.")